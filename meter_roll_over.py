import openpyxl
import PySimpleGUI as sg
from datetime import datetime
from actualizar import Actualizador

# Lista de rutas de archivo
rutas_archivos = [
    'C:\\Users\\slotr\\OneDrive\\Desktop\\slots_ariel\\meter_roll_over_soft2023.xlsx',
    '\\\\SLOTS-AC\\slots_ariel\\meter_roll_over_soft2023.xlsx'
]

# Ruta de archivo actual
ruta_archivo = None

# Libro de trabajo actual
workbook = None

# Crear la lista de opciones para la lista desplegable
opciones_supervisor = ["Ariel", "Ernest"]
opciones_meters = ["total in", "total out"]

# Crear el diseño de la ventana
layout = [
    [sg.Text("Lockname:"), sg.Input(key="-LOCKNAME-", size=(17, 1))],
    [sg.Text("Path:"), sg.Combo(rutas_archivos, key="-RUTA-", size=(50, 1))],
    [sg.Text("Meter:"), sg.Combo(opciones_meters, key="-METER-", size=(19, 1))],
    [sg.Text("Time:"), sg.Input(key="-TIME-", size=(17, 1))],
    [sg.Text("Manager:"), sg.Combo(opciones_supervisor, key="-SUPERVISOR-", size=(18, 1))],
    [sg.Button("Save")],
    [sg.Button("Check for Updates")],
    [sg.ProgressBar(100, orientation='h', size=(20, 20), key='-PROGRESS-')]
]

# Crear la ventana
ventana = sg.Window("RollOver", layout)
sg.theme('Light Blue 3')

# Ejecutar el bucle de eventos de la ventana
while True:
    evento, valores = ventana.read()

    if evento == sg.WINDOW_CLOSED:
        break
    elif evento == "Save":
        # Obtener lockname, meter, supervisor y tiempo del usuario
        ruta_archivo = valores["-RUTA-"]
        lockname = valores["-LOCKNAME-"]
        meter = valores["-METER-"]
        supervisor = valores["-SUPERVISOR-"]
        tiempo_str = valores["-TIME-"]

        # Si la ruta de archivo ha cambiado o no se ha abierto ningún archivo todavía
        if ruta_archivo != workbook.path if workbook else None or workbook is None:
            # Cerrar el archivo anterior si estaba abierto
            if workbook:
                workbook.close()

            # Intentar abrir el archivo existente o crear uno nuevo
            try:
                workbook = openpyxl.load_workbook(ruta_archivo)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

        # Obtener la fecha actual en el formato deseado
        fecha_actual = datetime.now().strftime("%Y-%m-%d")

        # Verificar si la hoja con la fecha actual ya existe en el archivo o si no hay ningún archivo abierto
        if workbook and fecha_actual not in workbook.sheetnames:
            # Crear una nueva hoja con el nombre de la fecha actual
            workbook.create_sheet(fecha_actual)

        # Obtener la hoja correspondiente a la fecha actual si hay un archivo abierto
        hoja = workbook[fecha_actual] if workbook else None

        # Obtener el número de fila siguiente para insertar los datos
        siguiente_fila = hoja.max_row + 1 if hoja else 1

        # Verificar si las columnas ya existen en la hoja
        if hoja and hoja['A1'].value is None:
            # Insertar nombre de columna "Lockname"
            hoja['A1'] = 'Lockname'
        if hoja and hoja['B1'].value is None:
            # Insertar nombre de columna "Meter"
            hoja['B1'] = 'Meter'
        if hoja and hoja['C1'].value is None:
            # Insertar nombre de columna "Date"
            hoja['C1'] = 'Date'
        if hoja and hoja['D1'].value is None:
            # Insertar nombre de columna "Time"
            hoja['D1'] = 'Time'
        if hoja and hoja['E1'].value is None:
            # Insertar nombre de columna "Manager"
            hoja['E1'] = 'Manager'

        # Convertir tiempo_str a objeto de tiempo
        tiempo = datetime.strptime(tiempo_str, "%H:%M").time()

        # Verificar si el lockname y meter ya existen en alguna fila
        if hoja:
            for fila in hoja.iter_rows(min_row=2, max_row=siguiente_fila - 1):
                if fila[0].value == lockname and fila[1].value == meter:
                    sg.popup("Lockname + Meter already exist.")
                    break
            else:
                # Insertar nueva fila con los datos
                hoja.cell(row=siguiente_fila, column=1, value=lockname)
                hoja.cell(row=siguiente_fila, column=2, value=meter)
                hoja.cell(row=siguiente_fila, column=3, value=datetime.now())
                hoja.cell(row=siguiente_fila, column=4, value=tiempo.strftime("%H:%M:%S"))
                hoja.cell(row=siguiente_fila, column=5, value=supervisor)

                # Guardar el archivo si hay un archivo abierto
                if workbook:
                    workbook.save(ruta_archivo)
                    sg.popup("Data saved. thanks ")
                else:
                    sg.popup("theres no files . thanks ")

        # Incrementar el número de fila siguiente si hay un archivo y una hoja abiertos
        if hoja:
            siguiente_fila += 1

    elif evento == "Check for Updates":
        progreso_barra = ventana['-PROGRESS-']
        progreso_barra.update(0)  # Restablecer la barra de progreso al inicio

        if Actualizador.hay_actualizacion_disponible():
            sg.popup("¡update avaliable!")
            progreso_barra.update(50)  # Actualizar la barra de progreso al 50%

            # Realizar la actualización
            Actualizador.realizar_actualizacion()
            progreso_barra.update(100)  # Actualizar la barra de progreso al 100%
            sg.popup("¡update succefully!")

        else:
            sg.popup("theres no update")

# Cerrar la ventana
ventana.close()
