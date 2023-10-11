import openpyxl
import PySimpleGUI as sg
from datetime import datetime
from actualizar import Actualizador
import subprocess

def meter_pro():
    workbook = None

    rutas_archivos = [
        'C:\\Users\\slotr\\OneDrive\\Desktop\\slots_ariel\\meter_roll_over_soft2023.xlsx',
        '\\\\SLOTS-AC\\slots_ariel\\meter_roll_over_soft2023.xlsx'
    ]

    layout = [
        [sg.Text("Lockname:"), sg.Input(key="-LOCKNAME-", size=(17, 1))],
        [sg.Text("Path:"), sg.Combo(rutas_archivos, key="-RUTA-", size=(50, 1))],
        [sg.Text("Meter:"), sg.Combo(["total in", "total out"], key="-METER-", size=(19, 1))],
        [sg.Text("Time:"), sg.Input(key="-TIME-", size=(17, 1))],
        [sg.Text("total_in_mach:"), sg.Input(key="-ROLL-", size=(17, 1))],
        [sg.Text("incrementation:"), sg.Input(key="-INC-", size=(17, 1))],
        [sg.Text("Result:"), sg.Text("", key="-RESULT-", size=(30, 1))],
        [sg.Button("Calculate")],
        [sg.Text("Manager:"), sg.Combo(["Ariel", "Ernest"], key="-SUPERVISOR-", size=(18, 1))],
        [sg.Button("Save")],
        [sg.Button("Check for Updates...")],
        [sg.ProgressBar(100, orientation='h', size=(20, 20), key='-PROGRESS-', visible=False)]
    ]

    ventana = sg.Window("Roll Over EGM", layout)
    sg.theme('Light Blue 3')

    while True:
        evento, valores = ventana.read()

        if evento == sg.WINDOW_CLOSED:
            break

        elif evento == "Calculate":
            total_in_mach = valores["-ROLL-"]
            incrementation = valores['-INC-']

            try:
                total_in_mach = int(total_in_mach)
                incrementation = int(incrementation)

                if total_in_mach <= 999999999 and incrementation <= 999999999:
                    resultado = (999999999 - total_in_mach) + incrementation
                else:
                    resultado = "Over 9 digits please type only 9 "

                ventana["-RESULT-"].update(f" {resultado}")

            except ValueError:
                sg.popup("Please enter a valid number in total_in_mach.")

        elif evento == "Save":
            ruta_archivo = valores["-RUTA-"]
            lockname = valores["-LOCKNAME-"]
            meter = valores["-METER-"]
            supervisor = valores["-SUPERVISOR-"]
            tiempo_str = valores["-TIME-"]
            total_in_mach = valores["-ROLL-"]
            incrementation = valores['-INC-']
            resultado= None

            if ruta_archivo != workbook.path if workbook else None or workbook is None:
                if workbook:
                    workbook.close()
                try:
                    workbook = openpyxl.load_workbook(ruta_archivo)
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()

            fecha_actual = datetime.now().strftime("%Y-%m-%d")

            if workbook and fecha_actual not in workbook.sheetnames:
                workbook.create_sheet(fecha_actual)

            hoja = workbook[fecha_actual] if workbook else None

            siguiente_fila = hoja.max_row + 1 if hoja else 1

            if hoja and hoja['A1'].value is None:
                hoja['A1'] = 'Lockname'
            if hoja and hoja['B1'].value is None:
                hoja['B1'] = 'Meter'
            if hoja and hoja['C1'].value is None:
                hoja['C1'] = 'Date'
            if hoja and hoja['D1'].value is None:
                hoja['D1'] = 'Time'
            if hoja and hoja['E1'].value is None:
                hoja['E1'] = 'Manager'

            tiempo = datetime.strptime(tiempo_str, "%H:%M").time()

            if hoja:
                for fila in hoja.iter_rows(min_row=2, max_row=siguiente_fila - 1):
                    if fila[0].value == lockname and fila[1].value == meter:
                        sg.popup("Lockname + Meter already exist.")
                        break
                else:
                    hoja.cell(row=siguiente_fila, column=1, value=lockname)
                    hoja.cell(row=siguiente_fila, column=2, value=meter)
                    hoja.cell(row=siguiente_fila, column=3, value=datetime.now())
                    hoja.cell(row=siguiente_fila, column=4, value=tiempo.strftime("%H:%M:%S"))
                    hoja.cell(row=siguiente_fila, column=5, value=supervisor)
                    hoja.cell(row=siguiente_fila, column=6, value=total_in_mach)
                    hoja.cell(row=siguiente_fila, column=7, value=incrementation)

                    hoja.cell(row=siguiente_fila, column=8, value=resultado)




                    if workbook:
                        workbook.save(ruta_archivo)
                        sg.popup("Data saved.... Thanks!")
                    else:
                        sg.popup("There are no files. Thanks!")

            if hoja:
                siguiente_fila += 1

        elif evento == "Check for Updates...":
            if Actualizador.hay_actualizacion_disponible():
                progreso_barra = ventana['-PROGRESS-']
                progreso_barra.UpdateBar(0, max=100)
                sg.popup("Updating...", no_titlebar=True, keep_on_top=True)

                try:
                    Actualizador.realizar_actualizacion(con_barra=True, barra=progreso_barra)
                    sg.popup("Update successful.")
                except subprocess.CalledProcessError as e:
                    sg.popup("Error updating: " + str(e))

                progreso_barra.UpdateBar(100)
            else:
                sg.popup("No updates available.")

    ventana.close()

meter_pro()
