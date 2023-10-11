import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from datetime import datetime

# Función para agregar una nueva máquina
def agregar_maquina():
    serial = entry_serial_var.get()
    brand = marca_var.get()
    lockname = entry_numero_maquina_var.get()
    modelo = modelo_var.get()

    # Validar que todos los campos estén llenos
    if serial == "" or brand == "" or lockname == "" or modelo == "":
        messagebox.showerror("Error", "Por favor, complete todos los campos.")
        return

    # Verificar que el lockname no esté en uso
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == lockname:
            messagebox.showerror("Error", "El Lockname ya está en uso.")
            return

    # Crear un objeto de máquina y agregarlo al conjunto
    maquina = Maquina(serial, brand, lockname, modelo)
    maquinas_set.add(maquina)

    # Agregar las cabeceras de columna si no existen
    if not sheet['A1'].value:
        sheet['A1'] = 'Serial'
        sheet['B1'] = 'Brand'
        sheet['C1'] = 'Lockname'
        sheet['D1'] = 'Modelo'
        sheet['E1'] = 'Monitor'
        sheet['F1'] = 'Bill Acceptor'
        sheet['G1'] = 'Touch Screen'
        sheet['H1'] = 'Printer'
        sheet['I1'] = 'Power Supply'
        sheet['J1'] = 'Last Maintenance Date'

    # Agregar los datos a una nueva fila
    nueva_fila = [serial, brand, lockname, modelo, '', '', '', '', '', '']
    sheet.append(nueva_fila)

    # Guardar los cambios en el archivo Excel
    wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')

    messagebox.showinfo("Éxito", "Máquina agregada correctamente.")

# Función para buscar una máquina por Lockname
def buscar_maquina():
    lockname = entry_buscar_lockname.get()
    resultados_text.delete(1.0, tk.END)  # Limpiar resultados anteriores

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == lockname:
            # Mostrar los datos en el widget de resultados
            resultado = f"Serial: {row[0]}\nMarca: {row[1]}\nLockname: {row[2]}\nModelo: {row[3]}\nMonitor: {row[4]}\nBill Acceptor: {row[5]}\nTouch Screen: {row[6]}\nPrinter: {row[7]}\nPower Supply: {row[8]}\nLast Maintenance Date: {row[9]}"
            resultados_text.insert(tk.END, resultado)
            return

    resultados_text.insert(tk.END, f"No se encontró ninguna máquina con el Lockname {lockname}.")

# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
# Función para agregar componentes a una máquina
def agregar_componente():
    lockname = entry_buscar_lockname.get()
    componente_seleccionado = componente_var.get()

    if not lockname:
        messagebox.showwarning("Advertencia", "Primero busque una máquina por Lockname.")
        return

    row_index = None

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] == lockname:
            row_index = index
            break

    if row_index is not None:
        componente_columna = columnas_componentes.get(componente_seleccionado)
        if componente_columna is not None:
            if sheet.cell(row=row_index, column=componente_columna).value == "Yes":
                messagebox.showinfo("Información", f"{componente_seleccionado} ya ha sido agregado a esta máquina.")
            else:
                sheet.cell(row=row_index, column=componente_columna).value = "Yes"
                sheet.cell(row=row_index, column=10).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')
                messagebox.showinfo("Éxito", f"{componente_seleccionado} agregado correctamente.")
        else:
            messagebox.showwarning("Advertencia", f"No se reconoce el componente: {componente_seleccionado}.")
    else:
        messagebox.showwarning("Advertencia", f"No se encontró ninguna máquina con el Lockname {lockname}.")





def realizar_mantenimiento():
    opcion = mantenimiento_var.get()
    lockname = entry_buscar_lockname.get()

    if opcion and lockname:
        row_index = None  # Inicializamos row_index como None

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[2] == lockname:
                row_index = index  # Almacenamos el índice de la fila
                break  # Salimos del bucle una vez que encontramos la fila

        if row_index is not None:
            # Realizar el mantenimiento o servicio completo
            if opcion == "Service":
                for componente in columnas_componentes:
                    sheet.cell(row=row_index, column=columnas_componentes[componente]).value = ""
                sheet.cell(row=row_index, column=10).value = ""
                wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')
                messagebox.showinfo("Éxito", "Servicio completo realizado con éxito.")
            else:
                sheet.cell(row=row_index, column=columnas_componentes[opcion]).value = ""
                wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')
                messagebox.showinfo("Éxito", f"Mantenimiento de {opcion} realizado con éxito.")
            actualizar_servicio_completo()
        else:
            messagebox.showwarning("Advertencia", f"No se encontró ninguna máquina con el Lockname {lockname}.")
    else:
        messagebox.showwarning("Advertencia", "Primero busque una máquina por Lockname y seleccione una opción de mantenimiento.")

def mostrar_maquinas():
    if not maquinas_set:
        messagebox.showinfo("Información", "No hay máquinas registradas.")
        return

    maquinas_info = "Máquinas Registradas:\n\n"
    for maquina in maquinas_set:
        maquinas_info += f"Serial: {maquina.serial}\n"
        maquinas_info += f"Marca: {maquina.brand}\n"
        maquinas_info += f"Lockname: {maquina.lockname}\n"
        maquinas_info += f"Modelo: {maquina.modelo}\n\n"

    messagebox.showinfo("Máquinas Registradas", maquinas_info)

def actualizar_servicio_completo():
    lockname = entry_buscar_lockname.get()
    if lockname:
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[2] == lockname:
                row_index = index
                components = row[4:9]
                if all(component == "Yes" for component in components):
                    sheet.cell(row=row_index, column=10).value = "Yes"
                else:
                    sheet.cell(row=row_index, column=10).value = ""
                wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')

class Maquina:
    def __init__(self, serial, brand, lockname, modelo):
        self.serial = serial
        self.brand = brand
        self.lockname = lockname
        self.modelo = modelo

# Abrir el archivo Excel
wb = openpyxl.load_workbook('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')
sheet = wb.active

# Conjunto para almacenar objetos de máquinas
maquinas_set = set()

# Diccionario para las marcas y modelos
marcas_modelos = {
    "Egt": ["Slam top", "Upright", "Roullette"],
    "Novomatic": ["Slam top", "Upright"],
    "Igt": ["Slam top", "Upright"],
    "Alfa street": ["Roullette"]
}

# Diccionario para las columnas de componentes
columnas_componentes = {
    "Monitor": 5,
    "Bill Acceptor (JCM)": 6,
    "Bill Acceptor (MEI)": 6,
    "Touch Screen (Ithaca)": 7,
    "Touch Screen (Gen2)": 7,
    "Printer (Ithaca)": 8,
    "Printer (Gen2)": 8,
    "Power Supply (12V)": 9,
    "Power Supply (24V)": 9,
}

# Crear la ventana principal
window = tk.Tk()
window.title("Machine List")
window.geometry("400x600")  # Ajusta el tamaño de la ventana

# Etiquetas y campos de entrada para agregar máquinas
label_serial = tk.Label(window, text="Serial:")
entry_serial_var = tk.StringVar()
marca_var = tk.StringVar()
entry_numero_maquina_var = tk.StringVar()
modelo_var = tk.StringVar()

label_serial = tk.Label(window, text="Serial:")
label_marca = tk.Label(window, text="Marca:")
label_numero_maquina = tk.Label(window, text="Lockname:")
label_modelo = tk.Label(window, text='Modelo')

entry_serial = tk.Entry(window, textvariable=entry_serial_var)
entry_numero_maquina = tk.Entry(window, textvariable=entry_numero_maquina_var)

marca_combobox = ttk.Combobox(window, textvariable=marca_var, values=list(marcas_modelos.keys()))
marca_combobox.grid(row=1, column=1)

label_modelo.grid(row=3, column=0)
modelo_combobox = ttk.Combobox(window, textvariable=modelo_var, values=[])
modelo_combobox.grid(row=3, column=1)
button_mostrar_maquinas = tk.Button(window, text="Mostrar Máquinas", command=mostrar_maquinas)
button_mostrar_maquinas.grid(row=13, column=0, columnspan=2)
# Función para actualizar los modelos cuando se selecciona una marca
def actualizar_modelos(event):
    selected_marca = marca_var.get()
    modelos = marcas_modelos.get(selected_marca, [])
    modelo_combobox["values"] = modelos
    modelo_var.set("")  # Reiniciar el valor del modelo

marca_combobox.bind("<<ComboboxSelected>>", actualizar_modelos)

button_agregar = tk.Button(window, text="ADD MACHINE", command=agregar_maquina)

# Etiquetas y campos de entrada para buscar máquinas por Lockname
label_buscar_lockname = tk.Label(window, text="SEARCH by Lockname:")
entry_buscar_lockname = tk.StringVar()
entry_buscar_lockname = tk.Entry(window, textvariable=entry_buscar_lockname)
button_buscar_lockname = tk.Button(window, text="SEARCH by Lockname", command=buscar_maquina)

# Botón para agregar componentes y realizar mantenimiento
label_componente = tk.Label(window, text="Componente:")
componente_var = tk.StringVar()
componente_combobox = ttk.Combobox(window, textvariable=componente_var, values=list(columnas_componentes.keys()))
button_agregar_componente = tk.Button(window, text="Agregar Componente", command=agregar_componente)



# Botones para realizar mantenimiento
mantenimiento_var = tk.StringVar()
mantenimiento_combobox = ttk.Combobox(window, textvariable=mantenimiento_var, values=["Service"] + list(columnas_componentes.keys()))
button_realizar_mantenimiento = tk.Button(window, text="Realizar Mantenimiento", command=realizar_mantenimiento)

# Widget de resultados
resultados_text = tk.Text(window, height=10, width=40)
resultados_text.grid(row=12, column=0, columnspan=3)

# Organizar elementos en la ventana
label_serial.grid(row=0, column=0)
label_marca.grid(row=1, column=0)
label_numero_maquina.grid(row=2, column=0)
label_modelo.grid(row=3, column=0)

entry_serial.grid(row=0, column=1)
entry_numero_maquina.grid(row=2, column=1)

button_agregar.grid(row=6, column=0, columnspan=2)

label_buscar_lockname.grid(row=7, column=0)
entry_buscar_lockname.grid(row=7, column=1)
button_buscar_lockname.grid(row=7, column=2)

# Organizar botones para agregar componentes y realizar mantenimiento
label_componente.grid(row=8, column=0)
componente_combobox.grid(row=8, column=1)
button_agregar_componente.grid(row=8, column=2)

mantenimiento_combobox.grid(row=9, column=1)
button_realizar_mantenimiento.grid(row=9, column=2)

# Widget de resultados
resultados_text = tk.Text(window, height=10, width=40)
resultados_text.grid(row=12, column=0, columnspan=3)

# Iniciar la ventana
window.mainloop()
