import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

# Función para agregar una nueva máquina
def agregar_maquina():
    serial = entry_serial_var.get()
    brand = marca_var.get()
    lockname = entry_numero_maquina_var.get()
    modelo = modelo_var.get()

    # Validar que todos los campos estén llenos
    if serial == "" or brand == "" or lockname == "" or modelo == "":
        messagebox.showerror("Error", "Por favor, complete todos los campos.")
        return

    # Verificar que el lockname no esté en uso
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == lockname:
            messagebox.showerror("Error", "El Lockname ya está en uso.")
            return

    # Crear un objeto de máquina y agregarlo al conjunto
    maquina = (serial, brand, lockname, modelo)
    maquinas_set.add(maquina)

    # Agregar las cabeceras de columna si no existen
    if not sheet['A1'].value:
        sheet['A1'] = 'Serial'
        sheet['B1'] = 'Brand'
        sheet['C1'] = 'Lockname'
        sheet['D1'] = 'Modelo'

    # Agregar los datos a una nueva fila
    nueva_fila = [serial, brand, lockname, modelo]
    sheet.append(nueva_fila)

    # Guardar los cambios en el archivo Excel
    wb.save('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')

    messagebox.showinfo("Éxito", "Máquina agregada correctamente.")

# Función para buscar una máquina por lockname
def buscar_maquina():
    lockname = entry_buscar_lockname.get()
    resultados_text.delete(1.0, tk.END)  # Limpiar resultados anteriores

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == lockname:
            # Mostrar los datos en el widget de resultados
            resultado = f"Serial: {row[0]}\nMarca: {row[1]}\nLockname: {row[2]}\nModelo: {row[3]}"
            resultados_text.insert(tk.END, resultado)
            return

    resultados_text.insert(tk.END, f"No se encontró ninguna máquina con el Lockname {lockname}.")

# Abrir el archivo Excel
wb = openpyxl.load_workbook('C:\\Users\\slotr\\Desktop\\tragamonedas.xlsx')
sheet = wb.active

# Conjunto para almacenar objetos de máquinas
maquinas_set = set()

# Diccionario para las marcas y modelos
marcas_modelos = {
    "Egt": ["Slam top", "Upright", "Roullette"],
    "Novomatic": ["Slam top", "Upright"],
    "Igt": ["Slam top", "Upright"],
    "Alfa street": ["Roullette"]
}


# Crear la ventana principal
window = tk.Tk()
window.title("Machine List")
window.geometry("400x400")  # Ajusta el tamaño de la ventana

# Etiquetas y campos de entrada para agregar máquinas
label_serial = tk.Label(window, text="Serial:")
entry_serial_var = tk.StringVar()
marca_var = tk.StringVar()
entry_numero_maquina_var = tk.StringVar()
modelo_var = tk.StringVar()

label_serial = tk.Label(window, text="Serial:")
label_marca = tk.Label(window, text="Marca:")
label_numero_maquina = tk.Label(window, text="Lockname:")
label_modelo = tk.Label(window, text='Modelo')

entry_serial = tk.Entry(window, textvariable=entry_serial_var)
entry_numero_maquina = tk.Entry(window, textvariable=entry_numero_maquina_var)

# Listas desplegables para marcas y modelos
label_marca.grid(row=1, column=0)
marca_combobox = ttk.Combobox(window, textvariable=marca_var, values=list(marcas_modelos.keys()))
marca_combobox.grid(row=1, column=1)

label_modelo.grid(row=3, column=0)
modelo_combobox = ttk.Combobox(window, textvariable=modelo_var, values=[])
modelo_combobox.grid(row=3, column=1)

# Función para actualizar los modelos cuando se selecciona una marca
def actualizar_modelos(event):
    selected_marca = marca_var.get()
    modelos = marcas_modelos.get(selected_marca, [])
    modelo_combobox["values"] = modelos
    modelo_var.set("")  # Reiniciar el valor del modelo

marca_combobox.bind("<<ComboboxSelected>>", actualizar_modelos)

button_agregar = tk.Button(window, text="ADD MACHINE", command=agregar_maquina)

# Etiquetas y campos de entrada para buscar máquinas por Lockname
label_buscar_lockname = tk.Label(window, text="SEARCH by Lockname:")
entry_buscar_lockname = tk.StringVar()
entry_buscar_lockname = tk.Entry(window, textvariable=entry_buscar_lockname)
button_buscar_lockname = tk.Button(window, text="SEARCH by Lockname", command=buscar_maquina)

# Widget de resultados
resultados_text = tk.Text(window, height=10, width=40)
resultados_text.grid(row=8, column=0, columnspan=3)

# Organizar elementos en la ventana
label_serial.grid(row=0, column=0)
label_marca.grid(row=1, column=0)
label_numero_maquina.grid(row=2, column=0)
label_modelo.grid(row=3, column=0)

entry_serial.grid(row=0, column=1)
entry_numero_maquina.grid(row=2, column=1)

button_agregar.grid(row=6, column=0, columnspan=2)

label_buscar_lockname.grid(row=7, column=0)
entry_buscar_lockname.grid(row=7, column=1)
button_buscar_lockname.grid(row=7, column=2)

# Iniciar la ventana
window.mainloop()
